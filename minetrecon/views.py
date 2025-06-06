import io
import pandas as pd
from django.http import HttpResponse
from .models import Accounts, Users,Uploads,Audit  # Assuming this is the model for account mapping
from openpyxl import load_workbook
from .forms import MultiFileUploadForm,DocumentDownloadForm
from django.shortcuts import render,redirect
from django.http import JsonResponse
import requests
from openpyxl import load_workbook
import requests
from django.contrib.auth import logout
import os
from django.conf import settings
from datetime import datetime
from zipfile import ZipFile
import logging
from django.utils import timezone
import pytz
from django.utils.timezone import now
import pytz




def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
             # Log empty field attempt
            Audit.objects.create(
                username=username,
                user_action="Login failed: Empty fields",
                date_time=timezone.now(),
                upload=None
            )
            return JsonResponse({'error': 'Fields cannot be empty'}, status=400)

        login_url = f'https://ussd.minet.co.ke/api/login.php?username={username}&password={password}'
        response = requests.get(login_url)
        data = response.json()

        if data['status'] == 0:  # Assuming '0' means success
            user_data = data['data'][0]
            request.session['user_data'] = user_data  # Save user data in session
            request.session['username'] = username    # Save the entered username

            # Check if the username exists in the database
            if not Users.objects.filter(username=username).exists():
                # Add the username if it doesn't already exist
                Users.objects.create(username=username)

            # Log successful login attempt
            Audit.objects.create(
                username=username,
                user_action="Login successful",
                date_time=timezone.now().astimezone(pytz.timezone('Africa/Nairobi')),
                upload=None
            )
            return JsonResponse({'redirect_url': '/upload/'})  # Redirect on success
        else:
            error_message = data.get('message', 'Unknown error')

            # Log unsuccessful login attempt
            Audit.objects.create(
                username=username,
                user_action=f"Login failed: {error_message}",
                date_time=timezone.now().astimezone(pytz.timezone('Africa/Nairobi')),
                upload=None
            )

            return JsonResponse({'error': 'Login failed. ' + data.get('message', 'Unknown error')}, status=400)

    return render(request, 'login.html')

def logout_view(request):
    username = request.session.get('username', 'Unknown User')  # Retrieve username from session
    # Log the logout action in the Audit table
    Audit.objects.create(
        username=username,
        user_action="Logout successful",
        date_time=now().astimezone(pytz.timezone('Africa/Nairobi')),
        upload=None  # Assuming no file upload is associated with logout
    )
    logout(request)
    return redirect('login')  # Redirect to login page

logger = logging.getLogger(__name__)


def history_view(request):
    form = DocumentDownloadForm()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and 'date' in request.GET:
        # Handle AJAX request for fetching documents
        selected_date = request.GET.get('date')
        username = request.session.get('username')
        try:
            parsed_date = datetime.strptime(selected_date, '%d/%m/%Y').date()
            uploads = Uploads.objects.filter(username=username, date_time__date=parsed_date)
            document_choices = []

            # Populate choices from available documents
            for upload in uploads:
                if upload.bank_statement:
                    document_choices.append((upload.bank_statement, "Bank Statement"))
                if upload.general_ledger:
                    document_choices.append((upload.general_ledger, "General Ledger"))
                if upload.recon_document:
                    document_choices.append((upload.recon_document, "Reconciliation Document"))

            return JsonResponse({'documents': document_choices})

        except ValueError:
            return JsonResponse({'error': "Invalid date format"}, status=400)

    elif request.method == 'POST':
        if 'download' in request.POST:
            selected_date = request.POST.get('date')
            username = request.session.get('username')

            if not selected_date:
                return render(request, 'history.html', {
                    'form': form,
                    'error': "Please select a date."
                })

            # Convert 'DD/MM/YYYY' format to 'YYYY-MM-DD'
            try:
                parsed_date = datetime.strptime(selected_date, '%d/%m/%Y').date()
            except ValueError:
                return render(request, 'history.html', {
                    'form': form,
                    'error': "Please select a valid date in 'DD/MM/YYYY' format."
                })

            # Filter uploads by username and date
            uploads = Uploads.objects.filter(
                username=username,
                date_time__date=parsed_date
            )

            if not uploads:
                return render(request, 'history.html', {
                    'form': form,
                    'error': "No documents found for the selected date."
                })

            file_paths = []
            for upload in uploads:
                if upload.bank_statement:
                    bank_statement_path = os.path.join(settings.UPLOADS_DIR, upload.bank_statement)
                    if os.path.exists(bank_statement_path):
                        file_paths.append(bank_statement_path)

                if upload.general_ledger:
                    general_ledger_path = os.path.join(settings.UPLOADS_DIR, upload.general_ledger)
                    if os.path.exists(general_ledger_path):
                        file_paths.append(general_ledger_path)

                if upload.recon_document:
                    recon_document_path = os.path.join(settings.UPLOADS_RECONCILED_DIR, upload.recon_document)
                    if os.path.exists(recon_document_path):
                        file_paths.append(recon_document_path)

            if not file_paths:
                return render(request, 'history.html', {
                    'form': form,
                    'error': "No valid files found for the selected date."
                })

            # Create and serve zip file
            zip_buffer = io.BytesIO()
            with ZipFile(zip_buffer, 'w') as zip_file:
                for file_path in file_paths:
                    zip_file.write(file_path, os.path.basename(file_path))

            zip_buffer.seek(0)
            
            # Log the ZIP download in the Audit table
            Audit.objects.create(
                username=username,
                user_action="ZIP downloaded",
                date_time=now().astimezone(pytz.timezone('Africa/Nairobi')),
                upload=None  # No specific file upload related to this action
            )


            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=excel_documents.zip'
            return response

    return render(request, 'history.html', {'form': form})

# JSON data endpoint
def view_documents(request):
    selected_date = request.GET.get('date')
    if not selected_date:
        return JsonResponse({"error": "No date provided"}, status=400)

    try:
        date_obj = datetime.strptime(selected_date, "%d/%m/%Y").date()
    except ValueError:
        return JsonResponse({"error": "Invalid date format"}, status=400)

    documents = Uploads.objects.filter(
        date_time__date=date_obj,
        username=request.session.get('username')
    ).values('bank_statement', 'general_ledger', 'recon_document')

    doc_data = [
        {
            'bank_statement': doc.get('bank_statement'),
            'general_ledger': doc.get('general_ledger'),
            'recon_document': doc.get('recon_document')
        }
        for doc in documents
    ]
    return render(request, 'view_documents.html', {'documents': doc_data})

def filter_by_month_year(bank_df, gl_df, selected_month, selected_year):
    # Convert selected month and year to integers
    selected_month = int(selected_month)
    selected_year = int(selected_year)

    # Print input DataFrames for debugging
    #print("Bank DataFrame before filtering:")
    #print(bank_df.head())
    #print("General Ledger DataFrame before filtering:")
    #print(gl_df.head())

    # Filter Bank Statement based on 'Account Name:'
    filtered_bank_df = bank_df[bank_df['Account Name:'].notna()].copy()
    filtered_bank_df['Date'] = pd.to_datetime(filtered_bank_df['Account Name:'], errors='coerce')

    # Filter General Ledger based on 'Unnamed: 1'
    filtered_gl_df = gl_df[gl_df['Unnamed: 1'].notna()].copy()

    # Extract relevant dates from column B starting below the 6th cell (row index 5)
    filtered_gl_df = filtered_gl_df.iloc[6:].copy()  # Start from row index 6
    filtered_gl_df['Date'] = filtered_gl_df['Unnamed: 1'].str.extract(r'(\d{2}/\d{2}/\d{4})')[0]  # Extract date (DD/MM/YYYY)
    filtered_gl_df['Period'] = filtered_gl_df['Unnamed: 1'].str.extract(r'(\d{2} \d{4})')[0]  # Extract period (MM YYYY)

    # Combine Period and Date into a single datetime column
    filtered_gl_df['Parsed Date'] = pd.to_datetime(filtered_gl_df['Date'], format='%d/%m/%Y', errors='coerce')
    filtered_gl_df['Parsed Period'] = pd.to_datetime(filtered_gl_df['Period'], format='%m %Y', errors='coerce')

    # Print parsed dates for debugging
    #print("Parsed Dates and Periods in General Ledger DataFrame:")
    #print(filtered_gl_df[['Unnamed: 1', 'Parsed Date', 'Parsed Period']].head(20))  # Show the parsed dates

    # Apply filtering based on selected month and year
    filtered_bank_df = filtered_bank_df[(filtered_bank_df['Date'].dt.month == selected_month) & 
                                        (filtered_bank_df['Date'].dt.year == selected_year)]

    # Filter General Ledger based on the parsed period
    filtered_gl_df = filtered_gl_df[(filtered_gl_df['Parsed Period'].dt.month == selected_month) & 
                                     (filtered_gl_df['Parsed Period'].dt.year == selected_year)]


    

    # Print the results of filtering for debugging
    #print("Filtered Bank DataFrame:")
    #print(filtered_bank_df)
    #print("Filtered General Ledger DataFrame:")
    #print(filtered_gl_df)


    
    # Check if any records were found after filtering
    if filtered_gl_df.empty:
        print("No matching records found in the General Ledger.")

    # Rename columns in the Bank Statement sheet
    filtered_bank_df.rename(columns={
        filtered_bank_df.columns[0]: 'Transaction Date',
        filtered_bank_df.columns[1]: 'Description',
        filtered_bank_df.columns[2]: 'Value Date',
        filtered_bank_df.columns[3]: 'Debit',
        filtered_bank_df.columns[4]: 'Credit',
        filtered_bank_df.columns[5]: 'Balance'
    }, inplace=True)

    # Drop the 'Date' column
    if 'Date' in filtered_bank_df.columns:
        filtered_bank_df.drop(columns=['Date'], inplace=True)

    filtered_bank_df.iloc[0, 6:] = None


    filtered_gl_df.rename(columns={
        filtered_gl_df.columns[0]: 'Batch #',
        filtered_gl_df.columns[1]: 'Period/Date',
        filtered_gl_df.columns[2]: 'Description',
        filtered_gl_df.columns[3]: 'BRT',
        filtered_gl_df.columns[4]: 'BRun',
        filtered_gl_df.columns[5]: 'SRC',
        filtered_gl_df.columns[6]: 'Orig Currency',
        filtered_gl_df.columns[7]: 'Currency',
        filtered_gl_df.columns[8]: 'Rate',
        filtered_gl_df.columns[9]: 'Amount'
        
    }, inplace=True)

    # Drop the 'Date' column
    if 'Unamed: 6' in filtered_gl_df.columns:
        filtered_gl_df.drop(columns=['Unamed: 6'], inplace=True)

    if 'Date' in filtered_gl_df.columns:
        filtered_gl_df.drop(columns=['Date'], inplace=True)

    if 'Period' in filtered_gl_df.columns:
        filtered_gl_df.drop(columns=['Period'], inplace=True)
    
    if 'Parsed Date' in filtered_gl_df.columns:
        filtered_gl_df.drop(columns=['Parsed Date'], inplace=True)

    if 'Parsed Period' in filtered_gl_df.columns:
        filtered_gl_df.drop(columns=['Parsed Period'], inplace=True)

    
   
    return filtered_bank_df.reset_index(drop=True), filtered_gl_df.reset_index(drop=True)







# Upload files and process them
def upload_file(request):
    if request.method == 'POST':
        form = MultiFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file1 = request.FILES['file1']  # Bank Statement Workbook
            file2 = request.FILES['file2']  # General Ledger


            

            # Define the upload path
            upload_dir = os.path.join(settings.BASE_DIR, 'uploads')
            reconciled_dir = os.path.join(settings.BASE_DIR, 'uploads_reconciled')
            os.makedirs(upload_dir, exist_ok=True)  # Create 'uploads' directory if it doesn't exist
            os.makedirs(reconciled_dir, exist_ok=True)
            

            # Get the current date and time as a string for appending to file names
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Prepare the filenames by replacing spaces with underscores and appending the timestamp
            file1_name = f"NCBA_BANK_STATEMENT_{timestamp}.xlsx"
            file2_name = f"GENERAL_LEDGER_{timestamp}.xlsx"
            reconciled_file_name = f"reconciled_data_{timestamp}.xlsx"

            # Save the files in the 'uploads' directory
            file1_path = os.path.join(upload_dir, file1_name)
            file2_path = os.path.join(upload_dir, file2_name)
            reconciled_file_path = os.path.join(reconciled_dir, reconciled_file_name)


           
            try:
               
                
                # Step 1: Extract GLID from the General Ledger


                df2 = pd.read_excel(file2)
              
                processed_df2 = df2.copy() # copy for manipulation the general ledger
                if len(df2) >= 5:
                    account_number_gl = df2.iloc[4, 0]  # Extract GLID from row 6, column A
                else:
                    return HttpResponse("No account GLID found in the General Ledger.")

                # Step 2: Load the Bank Statement Workbook and search for the bank account number
                wb = load_workbook(file1)
                matched_sheet = None
                account_number_bank = None

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    cell_value = ws['B3'].value  # Replace 'B6' with the actual location of the bank account
                    if cell_value:
                        account_number_bank = cell_value
                        # Step 3: Check if GLID and account number are mapped in the database
                        mapping = Accounts.objects.filter(glaccount_number=account_number_gl, bankaccount_number=account_number_bank).first()

                        if mapping:
                            matched_sheet = sheet  # Keep track of the matching sheet
                            break  # Exit after finding the match

                if not matched_sheet:
                    username = request.session.get('username', 'guest')
                    Audit.objects.create(
                        username=username,
                        user_action="Upload failed: No matching account found between GLID and bank account",
                        date_time=timezone.now().astimezone(pytz.timezone('Africa/Nairobi')),
                        upload=None
                    )
                    return HttpResponse(f"No sheet found in the Bank Statement with account number {account_number_bank}.")
                
                
                sheet = pd.read_excel(file1, sheet_name=matched_sheet)
                
                processed_df1 = sheet.copy()  # Copy for manipulation bank statement data
                #print(processed_df1.columns)

                
               

                    

               

                # Step 4: Filter by Month and Year
                selected_month = form.cleaned_data['month']
                selected_year = form.cleaned_data['year']

               
                # Assuming 'Account Number:' is in the first column (0-indexed) in the bank statement
                filtered_bank_df, filtered_gl_df = filter_by_month_year(processed_df1, processed_df2, selected_month, selected_year)
                

                
               
                # Extract relevant columns and filter out empty rows
                bank_debits_filtered = sheet.iloc[6:, [0, 1, 3]].dropna().reset_index(drop=True)
                bank_credits_filtered = sheet.iloc[6:, [0, 1, 4]].dropna().reset_index(drop=True)

                bank_debits_filtered.columns = ['Date', 'Transaction details', 'Debit amount']
                bank_credits_filtered.columns = ['Date', 'Transaction details', 'Credit amount']



                # Step 4: Process the General Ledger data
                gl_data = df2.iloc[5:, [1, 2, 9]].copy()
                gl_data.columns = ['Date', 'Narrative', 'Amount']
                gl_data['Amount'] = pd.to_numeric(gl_data['Amount'], errors='coerce')

                

              
                # Filter GL data by month (dates in column A)

                gl_debits = gl_data[gl_data['Amount'] > 0].copy()
                gl_credits = gl_data[gl_data['Amount'] < 0].copy()
                gl_credits['Amount'] = gl_credits['Amount'].abs()

                 # Rename 'Amount' to 'Debit amount' or 'Credit amount'
                gl_debits.rename(columns={'Amount': 'Debit amount'}, inplace=True)
                gl_credits.rename(columns={'Amount': 'Credit amount'}, inplace=True)

            # Add additional columns with default values
                 
                for df in [bank_debits_filtered, bank_credits_filtered, gl_debits, gl_credits]:
                    
                    df['Is Reconciled']='' 
                    df['Reconciliation Method']=''  
                    df['Reconciliation Reference'] = ''
                    df['Has duplicate'] = ''



               

            # Function to mark duplicates
                def mark_duplicates(df, amount_column):
                    duplicates = df[amount_column].duplicated(keep=False)
                    df['Has duplicate'] = duplicates.apply(lambda x: 'TRUE' if x else 'FALSE')

            # Mark duplicates for Bank Debits and Bank Credits
                mark_duplicates(bank_debits_filtered, 'Debit amount')
                mark_duplicates(bank_credits_filtered, 'Credit amount')

            # Mark duplicates for GL Debits and GL Credits
                mark_duplicates(gl_debits, 'Debit amount')
                mark_duplicates(gl_credits, 'Credit amount')

                bank_charges = bank_debits_filtered[
                    bank_debits_filtered['Transaction details'].str.contains(
                        'Transaction Charge|Excise Duty|Ledger fee|Witholding Tax|Transactional Fee |IB Bulk Transfer Charge|Guarantee Commission|Gaurantee cancellation commission', na=False
                    )
                ].copy()

                bank_charges['Debit amount'] = bank_charges['Debit amount']
                total_bank_charges_debit = bank_charges['Debit amount'].sum()

                
                
                # Step 2: Remove 'Is Reconciled' and 'Reconciliation Method' columns from bank charges
                bank_charges.drop(columns=['Is Reconciled', 'Reconciliation Method','Reconciliation Reference'], inplace=True)
                
                #Loop through the bank charges and add reference numbers from the matched rows
                for idx, row in bank_charges.iterrows():
                    transaction_detail = row['Transaction details']
                    reference_found = False
                    for sheet_idx in range(len(sheet)):
                         # Iterate through the sheet to find the matching 'Transaction details'
                        if sheet.iloc[sheet_idx, 1] == transaction_detail:  # Assuming 'Transaction details' are in column 1 (index 1)
                            if sheet_idx + 1 < len(sheet):
                                bank_charges.at[idx, 'Reference No'] = sheet.iloc[sheet_idx + 1, 1]  # Cell below
                            else:
                                bank_charges.at[idx, 'Reference No'] = 'N/A'  # No cell below
                            reference_found = True
                            break  # Exit loop once found
                    if not reference_found:
                        bank_charges.at[idx, 'Reference No'] = 'N/A'  # If no matching transaction detail found
                
                # Convert the 'Narrative' column to string type (handles both strings and numbers)
                gl_debits['Narrative'] = gl_debits['Narrative'].astype(str).str.strip()
                
                # Create a new sheet "Receipts" for GL Debits where Narrative contains 'Cash Receipts (BTS) Run'
                receipts_data = gl_debits[gl_debits['Narrative'].str.contains('Cash Receipts \\(BTS\\) ', na=False)].copy()
                receipts_data.drop(columns=['Is Reconciled', 'Reconciliation Method','Reconciliation Reference'], inplace=True)
                




                bank_debits_filtered['Reference No'] = ''
                bank_credits_filtered['Reference No'] = ''

                for bank_idx, bank_row in bank_debits_filtered.iterrows():
                    transaction_detail = bank_row['Transaction details']
                    reference_found = False

                    for sheet_idx in range(len(sheet)):
        # Iterate through the sheet to find the matching 'Transaction details'
                        if sheet.iloc[sheet_idx, 1] == transaction_detail:  # Assuming 'Transaction details' are in column 1 (index 1)
            # Transfer the value in the cell below the transaction detail to 'Reference No'
                            if sheet_idx + 1 < len(sheet):
                                bank_debits_filtered.at[bank_idx, 'Reference No'] = sheet.iloc[sheet_idx + 1, 1]  # Cell below
                            else:
                                bank_debits_filtered.at[bank_idx, 'Reference No'] = 'N/A'  # No cell below
                            reference_found = True
                            break  # Exit loop once found
                    if not reference_found:
                        bank_debits_filtered.at[bank_idx, 'Reference No'] = 'N/A'  # If no matching transaction detail found

                for bank_idx, bank_row in bank_credits_filtered.iterrows():
                    transaction_detail = bank_row['Transaction details']
                    reference_found = False
                    for sheet_idx in range(len(sheet)):
        # Iterate through the sheet to find the matching 'Transaction details'
                        if sheet.iloc[sheet_idx, 1] == transaction_detail:  # Assuming 'Transaction details' are in column 1 (index 1)
                            if sheet_idx + 1 < len(sheet):
                                bank_credits_filtered.at[bank_idx, 'Reference No'] = sheet.iloc[sheet_idx + 1, 1]  # Cell below
                            else:
                                bank_credits_filtered.at[bank_idx, 'Reference No'] = 'N/A'  # No cell below
                            reference_found = True
                            break  # Exit loop once found
                    if not reference_found:
                        bank_credits_filtered.at[bank_idx, 'Reference No'] = 'N/A'  # If no matching transaction detail found


                 
                    
            
            #for col in ['']:
                #bank_charges[col] = ''

            # Compare Debit amount in Bank Debits with Credit amount in GL Credits
                for bank_idx, bank_row in bank_debits_filtered.iterrows():
                    reconciled = False
                    for gl_idx, gl_row in gl_credits.iterrows():
                    # Check if both Debit amount and Credit amount are the same
                        if bank_row['Debit amount'] == gl_row['Credit amount']:
                        # Mark as reconciled with "BS Debit and GL Credit"
                            bank_debits_filtered.at[bank_idx, 'Is Reconciled'] = 'TRUE'
                            bank_debits_filtered.at[bank_idx, 'Reconciliation Method'] = 'BS Debit and GL Credit'
                            gl_credits.at[gl_idx, 'Is Reconciled'] = 'TRUE'
                            gl_credits.at[gl_idx, 'Reconciliation Method'] = 'BS Debit and GL Credit'
                            reconciled = True

                        if any(keyword in bank_row['Transaction details'] for keyword in ["Transaction Charge", "Excise Duty", "Ledger fee","Witholding Tax","Transactional Fee", "IB Bulk Transfer Charge", "Guarantee Commission", "Gaurantee cancellation commission"]):
            # Only update if the condition hasn't been set before
                            if bank_debits_filtered.at[bank_idx, 'Is Reconciled'] != 'TRUE' or \
                                bank_debits_filtered.at[bank_idx, 'Reconciliation Method'] != 'BS Debit and GL Credit':
                                bank_debits_filtered.at[bank_idx, 'Is Reconciled'] = 'TRUE'
                                bank_debits_filtered.at[bank_idx, 'Reconciliation Method'] = 'NCBA Bank Charges'
                                reconciled = True
                    if not reconciled:
                        bank_debits_filtered.at[bank_idx, 'Is Reconciled'] = 'FALSE'
                        bank_debits_filtered.at[bank_idx, 'Reconciliation Method'] = 'N/A'
                    

            

                for bank_idx, bank_row in bank_credits_filtered.iterrows():
                    reconciled = False
                    for gl_idx, gl_row in gl_debits.iterrows():
                    # Check if both Debit amount and Credit amount are the same
                        if bank_row['Credit amount'] == gl_row['Debit amount']:
                        # Mark as reconciled with "BS Credit and GL Debit"
                            bank_credits_filtered.at[bank_idx, 'Is Reconciled'] = 'TRUE'
                            bank_credits_filtered.at[bank_idx, 'Reconciliation Method'] = 'BS Credit and GL Debit'
                            gl_debits.at[gl_idx, 'Is Reconciled'] = 'TRUE'
                            gl_debits.at[gl_idx, 'Reconciliation Method'] = 'BS Credit and GL Debit'
                            reconciled = True
                    if not reconciled:
                        bank_credits_filtered.at[bank_idx, 'Is Reconciled'] = 'FALSE'
                        bank_credits_filtered.at[bank_idx, 'Reconciliation Method'] = 'N/A'
                    
                for df in [gl_debits, gl_credits]:
                    df['Is Reconciled'] = df['Is Reconciled'].replace('', 'FALSE')
                    df['Reconciliation Method'] = df['Reconciliation Method'].replace('', 'N/A')
                    
                    
                for df in [bank_debits_filtered, bank_credits_filtered, gl_debits, gl_credits]:
                    method_col = 'Reconciliation Method'
                    ref_col = 'Reconciliation Reference'
                    if 'Debit amount' in df.columns:
                        amount_col = 'Debit amount'
                    else:
                        amount_col = 'Credit amount'

                    for idx in df.index:
                        if df.at[idx, method_col] in ['BS Debit and GL Credit', 'BS Credit and GL Debit']:
                            df.at[idx, ref_col] = df.at[idx, amount_col]
                

                # Sum the 'Debit amount' where 'Is Reconciled' is 'FALSE' for Bank Debits and GL Debits
                unreconciled_bank_debits_sum = bank_debits_filtered[bank_debits_filtered['Is Reconciled'] == 'FALSE']['Debit amount'].sum()
                unreconciled_gl_debits_sum = gl_debits[gl_debits['Is Reconciled'] == 'FALSE']['Debit amount'].sum()

                unreconciled_bank_credits_sum = bank_credits_filtered[bank_credits_filtered['Is Reconciled'] == 'FALSE']['Credit amount'].sum()
                unreconciled_gl_credits_sum = gl_credits[gl_credits['Is Reconciled'] == 'FALSE']['Credit amount'].sum()
            
            # Count 'TRUE' and 'FALSE' values in the 'Is Reconciled' column for Bank Debits and GL Debits
                true_bank_debits_count = bank_debits_filtered['Is Reconciled'].value_counts().get('TRUE', 0)
                false_bank_debits_count = bank_debits_filtered['Is Reconciled'].value_counts().get('FALSE', 0)

                true_bank_credits_count = bank_credits_filtered['Is Reconciled'].value_counts().get('TRUE', 0)
                false_bank_credits_count = bank_credits_filtered['Is Reconciled'].value_counts().get('FALSE', 0)

                true_gl_debits_count = gl_debits['Is Reconciled'].value_counts().get('TRUE', 0)
                false_gl_debits_count = gl_debits['Is Reconciled'].value_counts().get('FALSE', 0)

                true_gl_credits_count = gl_credits['Is Reconciled'].value_counts().get('TRUE', 0)
                false_gl_credits_count = gl_credits['Is Reconciled'].value_counts().get('FALSE', 0)

                

                # Step 7: Output the results for download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl' ) as writer:

                    if not processed_df1.empty:
                        processed_df1.to_excel(writer, index=False, sheet_name='Bank Statement')
                    if not processed_df2.empty:
                        processed_df2.to_excel(writer, index=False, sheet_name='General Ledger')
                    if not filtered_bank_df.empty:
                        filtered_bank_df.to_excel(writer, index=False, sheet_name='Bank Statement edited')
                    if not filtered_gl_df.empty:
                        filtered_gl_df.to_excel(writer, index=False, sheet_name='General Ledger edited')
                    if not bank_debits_filtered.empty:
                        bank_debits_filtered.to_excel(writer, sheet_name='Bank Debits', index=False)
                    if not bank_credits_filtered.empty:
                        bank_credits_filtered.to_excel(writer, sheet_name='Bank Credits', index=False)
                    if not gl_debits.empty:
                        gl_debits.to_excel(writer, sheet_name='GL Debits', index=False)
                    if not gl_credits.empty:
                        gl_credits.to_excel(writer, sheet_name='GL Credits', index=False)
                    if not bank_charges.empty:
                        bank_charges.to_excel(writer, index=False, sheet_name='Bank Charges debited')
                    if not receipts_data.empty:
                        receipts_data.to_excel(writer, sheet_name='Receipts', index=False)


                    

                    wb = writer.book
                    
                    
                    bank_debits_sheet = wb['Bank Debits']
                    bank_credits_sheet = wb['Bank Credits']
                    gl_debits_sheet = wb['GL Debits']
                    gl_credits_sheet = wb['GL Credits']
                    bank_charges_sheet = wb['Bank Charges debited']

                    bank_debits_row_count = len(bank_debits_filtered) + 3
                    bank_debits_sheet.cell(row=bank_debits_row_count, column=6, value='Sum of unreconciled Debit amounts:')
                    bank_debits_sheet.cell(row=bank_debits_row_count + 1, column=6, value=unreconciled_bank_debits_sum)

                    # Similar process for Bank Credits
                    bank_credits_row_count = len(bank_credits_filtered) + 3
                    bank_credits_sheet.cell(row=bank_credits_row_count, column=6, value='Sum of unreconciled Credit amounts:')
                    bank_credits_sheet.cell(row=bank_credits_row_count + 1, column=6, value=unreconciled_bank_credits_sum)

    # Write the sum 3 rows below for GL Debits
                    gl_debits_row_count = len(gl_debits) + 3
                    gl_debits_sheet.cell(row=gl_debits_row_count, column=6, value='Unreconciled Debit amounts:')
                    gl_debits_sheet.cell(row=gl_debits_row_count + 1, column=6, value=unreconciled_gl_debits_sum)

    # Similar process for GL Credits
                    gl_credits_row_count = len(gl_credits) + 3
                    gl_credits_sheet.cell(row=gl_credits_row_count, column=6, value='Unreconciled Credit amounts:')
                    gl_credits_sheet.cell(row=gl_credits_row_count + 1, column=6, value=unreconciled_gl_credits_sum)

    # Write the total reconciled and unreconciled counts for Bank Debits
                    bank_debits_sheet.cell(row=bank_debits_row_count + 3, column=6, value='Total reconciled Debit entries:')
                    bank_debits_sheet.cell(row=bank_debits_row_count + 4, column=6, value=true_bank_debits_count)

                    bank_debits_sheet.cell(row=bank_debits_row_count + 6, column=6, value='Total unreconciled Debit entries:')
                    bank_debits_sheet.cell(row=bank_debits_row_count + 7, column=6, value=false_bank_debits_count)

    # Similar process for Bank Credits
                    bank_credits_sheet.cell(row=bank_credits_row_count + 3, column=6, value='Total reconciled Credit entries:')
                    bank_credits_sheet.cell(row=bank_credits_row_count + 4, column=6, value=true_bank_credits_count)

                    bank_credits_sheet.cell(row=bank_credits_row_count + 6, column=6, value='Total unreconciled Credit entries:')
                    bank_credits_sheet.cell(row=bank_credits_row_count + 7, column=6, value=false_bank_credits_count)

    # Similar process for GL Debits and Credits
                    gl_debits_sheet.cell(row=gl_debits_row_count + 3, column=6, value='Total reconciled GL Debit entries:')
                    gl_debits_sheet.cell(row=gl_debits_row_count + 4, column=6, value=true_gl_debits_count)

                    gl_debits_sheet.cell(row=gl_debits_row_count + 6, column=6, value='Total unreconciled GL Debit entries:')
                    gl_debits_sheet.cell(row=gl_debits_row_count + 7, column=6, value=false_gl_debits_count)

                    gl_credits_sheet.cell(row=gl_credits_row_count + 3, column=6, value='Total reconciled GL Credit entries:')
                    gl_credits_sheet.cell(row=gl_credits_row_count + 4, column=6, value=true_gl_credits_count)

                    gl_credits_sheet.cell(row=gl_credits_row_count + 6, column=6, value='Total unreconciled GL Credit entries:')
                    gl_credits_sheet.cell(row=gl_credits_row_count + 7, column=6, value=false_gl_credits_count)

                




                # Create recoonciliation summary sheet
                

                def calculate_percentage(true_count, false_count):
                        total_count = true_count + false_count
                        if total_count > 0:
                            percentage = (true_count / total_count) * 100
                        else:
                            percentage = 0
                        return percentage
                
                # Calculate percentages for each sheet
                bank_debits_percentage = calculate_percentage(true_bank_debits_count, false_bank_debits_count)
                bank_credits_percentage = calculate_percentage(true_bank_credits_count, false_bank_credits_count)
                gl_debits_percentage = calculate_percentage(true_gl_debits_count, false_gl_debits_count)
                gl_credits_percentage = calculate_percentage(true_gl_credits_count, false_gl_credits_count)


                # Write percentage reconciled entries to 'Reconciliation Summary' sheet
                output.seek(0)  # Reset the pointer to the start of the BytesIO stream
                
                wb = load_workbook(output)

                summary_sheet = wb.create_sheet(title='Reconciliation Summary')
                  # Ensure the sheet is visible
                summary_sheet.cell(row=1, column=1, value='Sheet Name')
                summary_sheet.cell(row=1, column=2, value='Reconciled Percentage')

                summary_sheet.cell(row=2, column=1, value='Bank Debits')
                summary_sheet.cell(row=2, column=2, value=f'{bank_debits_percentage:.2f}%')

                summary_sheet.cell(row=3, column=1, value='Bank Credits')
                summary_sheet.cell(row=3, column=2, value=f'{bank_credits_percentage:.2f}%')

                summary_sheet.cell(row=4, column=1, value='GL Debits')
                summary_sheet.cell(row=4, column=2, value=f'{gl_debits_percentage:.2f}%')

                summary_sheet.cell(row=5, column=1, value='GL Credits')
                summary_sheet.cell(row=5, column=2, value=f'{gl_credits_percentage:.2f}%')

                summary_sheet.sheet_state = 'visible'  # Ensure the sheet is visible

# Ensure you handle the case where there are no bank charges

# Summary sheet in general
# Writing unreconciled totals from Bank Debits, Bank Credits, GL Debits, GL Credits into Summary Sheet
# This code needs to be inserted within the `with pd.ExcelWriter` block after writing the individual sheets

# Write unreconciled debit and credit totals into the summary sheet
                
                  # Create Summary Sheet
                summary_sheet1 = wb.create_sheet(title='Summary')
                summary_sheet1.sheet_state = 'visible'  # Ensure the sheet is visible
                summary_sheet1.cell(row=3, column=2, value="BANK RECONCILIATION STATEMENT")
                summary_sheet1.cell(row=5, column=2, value="BANK NAME")
                summary_sheet1.cell(row=6, column=2, value="DIVISION NAME")
                summary_sheet1.cell(row=7, column=2, value="ACCOUNT NAME")
                summary_sheet1.cell(row=8, column=2, value="Position is at:")
                summary_sheet1.cell(row=10, column=2, value="BALANCE AS PER BANK STATEMENT")
                summary_sheet1.cell(row=10, column=7, value="")  # G10

# Add Section
                summary_sheet1.cell(row=12, column=2, value="Add:")
                summary_sheet1.cell(row=13, column=3, value="Debits in Bank not in GL")
                summary_sheet1.cell(row=14, column=3, value="Debits in GL not in Bank")
                summary_sheet1.cell(row=15, column=3, value="Bank Charges")
                summary_sheet1.cell(row=16, column=3, value="Bounced Customer Cheques")
                summary_sheet1.cell(row=17, column=3, value="Petty Difference")
                summary_sheet1.cell(row=18, column=3, value="Withholding tax")

# Use the previously calculated unreconciled sums from Bank Debits, Bank Credits, GL Debits, GL Credits
                summary_sheet1.cell(row=13, column=5, value=unreconciled_bank_debits_sum)  # Unreconciled debits from bank
                summary_sheet1.cell(row=14, column=5, value=unreconciled_gl_debits_sum)  # Unreconciled debits from GL
                summary_sheet1.cell(row=15, column=5, value=total_bank_charges_debit)  # Placeholder for Bank Charges
                summary_sheet1.cell(row=16, column=5, value="-")  # Placeholder for bounced cheques
                summary_sheet1.cell(row=17, column=5, value="-")  # Placeholder for petty differences
                summary_sheet1.cell(row=18, column=5, value="-")  # Placeholder for withholding tax
                summary_sheet1.cell(row=19, column=6, value="=SUM(E13:E18)")  # Total sum of unreconciled debits

# Less Section
                summary_sheet1.cell(row=20, column=2, value="Less:")
                summary_sheet1.cell(row=20, column=3, value="Credits in Bank not in GL")
                summary_sheet1.cell(row=21, column=3, value="Credits in GL not in Bank")
                summary_sheet1.cell(row=22, column=3, value="Interest Income")
                summary_sheet1.cell(row=23, column=3, value="Unrepresented Cheques")
                summary_sheet1.cell(row=24, column=3, value="Withholding Tax")

                summary_sheet1.cell(row=20, column=5, value=unreconciled_bank_credits_sum)  # Unreconciled credits from bank
                summary_sheet1.cell(row=21, column=5, value=unreconciled_gl_credits_sum)  # Unreconciled credits from GL
                summary_sheet1.cell(row=22, column=5, value="-")  # Placeholder for interest income
                summary_sheet1.cell(row=23, column=5, value="-")  # Placeholder for unrepresented cheques
                summary_sheet1.cell(row=24, column=5, value="-")  # Placeholder for withholding tax
                summary_sheet1.cell(row=25, column=6, value="=SUM(E20:E24)")  # Total sum of unreconciled credits

# Footer Section
                summary_sheet1.cell(row=28, column=4, value="Computed Balance")
                summary_sheet1.cell(row=28, column=7, value="=G10+F19-F25")  # Formula for computed balance

                summary_sheet1.cell(row=30, column=2, value="Balance as per Company books")
                summary_sheet1.cell(row=30, column=7, value="")  # Placeholder for balance as per company records

                summary_sheet1.cell(row=33, column=4, value="Reconciling Difference")
                summary_sheet1.cell(row=33, column=7, value="=G28-G30")  # Difference between computed and company balance

# Footer with preparer and reviewer
                summary_sheet1.cell(row=37, column=2, value="Prepared by:")
                summary_sheet1.cell(row=39, column=2, value="Reviewed by:")

                summary_sheet1.cell(row=37, column=4, value="Signature:")
                summary_sheet1.cell(row=39, column=4, value="Signature:")

                summary_sheet1.cell(row=37, column=6, value="Date:")
                summary_sheet1.cell(row=39, column=6, value="Date:")
                output = io.BytesIO()  # Reset the BytesIO object
                wb.save(output)
                #wb.save('reconciled_data.xlsx')
 
                #writer.sheets['Bank Charges debited'] = writer.book.add_worksheet('Bank Charges debited')
   
            # Set the cursor position to the beginning of the stream
                output.seek(0)

                # Save the reconciled workbook to `uploads_reconciled` folder
               
                with open(reconciled_file_path, 'wb') as reconciled_file:
                    reconciled_file.write(output.getvalue())

                # Verify if reconciled file has been successfully saved
                if not os.path.exists(reconciled_file_path):
                    Audit.objects.create(
                        username=username,
                        user_action="Upload failed: Reconciled file could not be saved",
                        date_time=timezone.now().astimezone(pytz.timezone('Africa/Nairobi')),
                        upload=None
                    )
                    return HttpResponse("Error: Reconciled file could not be saved.", status=500)


                 # Now proceed to save file1 and file2 only if reconciled file is saved
            
                with open(file1_path, 'wb') as f:
                    for chunk in file1.chunks():
                        f.write(chunk)

                with open(file2_path, 'wb') as f:
                    for chunk in file2.chunks():
                        f.write(chunk)

                # Save data to the database
                username = request.session.get('username', 'guest')
                
                  # get the username from the session
                upload_entry = Uploads.objects.create(
                    username=username,
                    bank_statement=file1_name,
                    general_ledger=file2_name,
                    recon_document=reconciled_file_name,
                    date_time=datetime.now().astimezone(pytz.timezone('Africa/Nairobi')),  # this will be automatically set with auto_now_add in models
                )

                # Record successful upload in the Audit table
                Audit.objects.create(
                    username=username,
                    user_action="Upload successful",
                    date_time=timezone.now().astimezone(pytz.timezone('Africa/Nairobi')),
                    upload=upload_entry
                )

            # Send the file to the user for download
               
                response= HttpResponse(open(reconciled_file_path, 'rb').read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=reconciled_data.xlsx'
                return response
       
            except Exception as e:
                
                
                return HttpResponse(f"An error occurred: {str(e)}")    
   
    else:
        form = MultiFileUploadForm()
    return render(request, 'index.html', {'form': form})
